import openpyxl

def save_to_excel(roll_number, name):
    workbook = openpyxl.Workbook()
    sheet = workbook.active

    # Check if the sheet already has data (this is for appending new data)
    has_data = (sheet.cell(1, 1).value is not None)

    if not has_data:
        # Set column headers if the sheet is empty
        sheet.cell(1, 1, "Roll Number")
        sheet.cell(1, 2, "Name")

    # Find the next empty row to add data
    row_number = sheet.max_row + 1 if has_data else 2
    sheet.cell(row_number, 1, roll_number)
    sheet.cell(row_number, 2, name)

    # Save the data to an Excel file
    workbook.save("student_data.xlsx")

if __name__ == "__main__":
    try:
        roll_number = int(input("Enter Roll Number: "))
        name = input("Enter Name: ")

        save_to_excel(roll_number, name)

        print("Data saved successfully to 'student_data.xlsx'.")
    except ValueError:
        print("Invalid input. Please enter a valid Roll Number.")
